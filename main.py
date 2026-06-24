"""
╔══════════════════════════════════════════════════════════╗
║         BRDrive - Agendador Principal                    ║
║  Executa a cada 25 minutos:                              ║
║  1. Extrai do JMS                                        ║
║  2. Processa dados                                       ║
║  3. Atualiza Database + BRDrive no OneDrive              ║
║  4. Dispara alertas via Faishu                           ║
╚══════════════════════════════════════════════════════════╝

USO:
    python main.py                  # roda continuamente a cada 25min
    python main.py --agora          # roda uma vez imediatamente
    python main.py --intervalo 30   # muda intervalo para 30 min
"""

import argparse
import logging
import os
import time
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from jms_extractor import extrair
from processador import (
    processar_database,
    processar_brdrive,
    detectar_datas_novas,
    gerar_alertas,
    DB_COLS,
)

# ── CONFIGURAÇÕES ─────────────────────────────────────────────────────────────
ONEDRIVE_BASE      = r"C:\Users\Robo Transporte\OneDrive - J&T EXPRESS - FILIAL SP"
ONEDRIVE_DIR       = os.getenv("ONEDRIVE_DIR", str(Path(ONEDRIVE_BASE) / "Power BI Gus"))
DATABASE_FILENAME  = "Database - Power bi.xlsx"
BRDRIVE_FILENAME   = "BRDrive_BI_Novo.xlsx"
LOG_DIR            = str(Path(__file__).parent / "logs")
INTERVALO_MINUTOS  = 25

DATABASE_PATH = str(Path(ONEDRIVE_DIR) / DATABASE_FILENAME)
BRDRIVE_PATH  = str(Path(ONEDRIVE_DIR) / BRDRIVE_FILENAME)

# ── LOGGING ───────────────────────────────────────────────────────────────────
Path(LOG_DIR).mkdir(parents=True, exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(
            f"{LOG_DIR}/main_{datetime.now().strftime('%Y%m%d')}.log",
            encoding="utf-8",
        ),
        logging.StreamHandler(),
    ],
)
log = logging.getLogger("BRDrive_Main")


# ── SALVAR EXCEL PRESERVANDO ABAS ─────────────────────────────────────────────
def salvar_fPrincipal(df: pd.DataFrame, caminho: str):
    """Salva o DataFrame na aba fPrincipal preservando as outras abas"""
    wb = load_workbook(caminho)
    ws = wb["fPrincipal"]

    # Limpar dados antigos (mantém cabeçalho)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    # Escrever novos dados
    for r_idx, row_data in enumerate(
        dataframe_to_rows(df, index=False, header=False), start=2
    ):
        for c_idx, value in enumerate(row_data, start=1):
            ws.cell(row=r_idx, column=c_idx).value = (
                None if (not isinstance(value, str) and pd.isna(value)) else value
            )

    wb.save(caminho)
    log.info(f"💾 Salvo: {Path(caminho).name}")


# ── CICLO PRINCIPAL ────────────────────────────────────────────────────────────
def executar_ciclo():
    log.info("━" * 55)
    log.info(f"🔄 CICLO INICIADO — {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    log.info("━" * 55)

    # ── 1. Verificar se os arquivos do OneDrive existem ───────────────
    if not Path(DATABASE_PATH).exists():
        log.error(f"❌ Database não encontrado: {DATABASE_PATH}")
        log.error("Configure a variável ONEDRIVE_DIR ou coloque os arquivos na pasta correta.")
        return
    if not Path(BRDRIVE_PATH).exists():
        log.error(f"❌ BRDrive não encontrado: {BRDRIVE_PATH}")
        return

    # ── 2. Extrair relatório do JMS ───────────────────────────────────
    log.info("📥 Extraindo dados do JMS...")
    arquivo_jms = extrair()
    if not arquivo_jms:
        log.error("❌ Extração do JMS falhou, pulando ciclo")
        return

    # ── 3. Ler arquivos ───────────────────────────────────────────────
    log.info("📂 Lendo arquivos...")
    try:
        df_jms = pd.read_excel(arquivo_jms)
        df_db  = pd.read_excel(DATABASE_PATH)
        df_br  = pd.read_excel(BRDRIVE_PATH, sheet_name="fPrincipal")
    except Exception as e:
        log.error(f"❌ Erro ao ler arquivos: {e}")
        return

    log.info(f"   JMS exportado: {len(df_jms)} linhas")
    log.info(f"   Database atual: {len(df_db)} linhas")
    log.info(f"   BRDrive atual: {len(df_br)} linhas")

    # ── 4. Detectar datas novas ───────────────────────────────────────
    datas_novas = detectar_datas_novas(df_db, df_br)
    if not datas_novas:
        log.info("✅ Nenhuma data nova para processar")
    else:
        # ── 5. Processar Database ─────────────────────────────────────
        log.info(f"⚙️  Processando Database para: {datas_novas}")
        df_db = processar_database(df_db, datas_novas)
        df_db.to_excel(DATABASE_PATH, index=False)
        log.info("💾 Database atualizado!")

        # ── 6. Atualizar BRDrive ──────────────────────────────────────
        log.info("⚙️  Atualizando BRDrive...")
        df_br_novo = processar_brdrive(df_br, df_db, datas_novas)
        salvar_fPrincipal(df_br_novo, BRDRIVE_PATH)
        log.info("💾 BRDrive atualizado!")

    # ── 7. Gerar e enviar alertas ─────────────────────────────────────
    hoje = datetime.now().strftime("%Y-%m-%d")
    alertas = gerar_alertas(df_db, hoje)

    if alertas:
        log.info(f"🚨 Enviando {len(alertas)} alertas...")
        try:
            from faishu_alertas import enviar_alertas
            enviar_alertas(alertas)
        except ImportError:
            log.warning("⚠️  Módulo faishu_alertas não encontrado — alertas não enviados")
            log.warning("    Cole o seu código Faishu em faishu_alertas.py")
        except Exception as e:
            log.error(f"❌ Erro ao enviar alertas: {e}")
    else:
        log.info("✅ Sem alertas para enviar")

    log.info(f"✅ CICLO CONCLUÍDO — {datetime.now().strftime('%H:%M:%S')}")


# ── LOOP CONTÍNUO ──────────────────────────────────────────────────────────────
def rodar_continuamente(intervalo_min: int):
    log.info("🚀 BRDrive Robot iniciado!")
    log.info(f"⏱️  Intervalo: {intervalo_min} minutos")
    log.info(f"📁 OneDrive: {ONEDRIVE_DIR}")
    log.info("Pressione CTRL+C para parar\n")

    while True:
        try:
            executar_ciclo()
        except KeyboardInterrupt:
            log.info("\n🛑 Robô encerrado pelo usuário")
            break
        except Exception as e:
            log.error(f"❌ Erro inesperado: {e}")

        proxima = intervalo_min * 60
        log.info(f"⏳ Próxima execução em {intervalo_min} minutos...\n")
        time.sleep(proxima)


# ── ENTRY POINT ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="BRDrive Robot")
    parser.add_argument("--agora",     action="store_true", help="Rodar uma vez agora")
    parser.add_argument("--intervalo", type=int, default=INTERVALO_MINUTOS,
                        help="Intervalo em minutos (padrão: 25)")
    args = parser.parse_args()

    if args.agora:
        executar_ciclo()
    else:
        rodar_continuamente(args.intervalo)
